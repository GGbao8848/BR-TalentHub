"""BR TalentHub 招聘会简历收集系统 —— FastAPI 后端。

第二版：
- 岗位管理：增删 + Excel 导入（识别"岗位名称/岗位要求"表头）
- 学校管理：多校并存，每校独立二维码，绑定岗位
- 简历按 学校/岗位/文件名 三级目录存储
- 数据看板：按学校/岗位/日期多维度统计
"""
import io
import os
import re
import socket
import uuid
import zipfile
from datetime import datetime
from io import BytesIO
from pathlib import Path

import qrcode
import uvicorn
from dotenv import load_dotenv
from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, Response
from fastapi.staticfiles import StaticFiles
from openpyxl import load_workbook

from app import database as db

BASE_DIR = Path(__file__).resolve().parent.parent
STATIC_DIR = BASE_DIR / "static"
DATA_DIR = BASE_DIR / "data"

# 加载项目根目录的 .env（HOST/PORT/SAVE_DIR 等配置）
load_dotenv(BASE_DIR / ".env")

# 监听地址与端口：从 .env 或环境变量读取（保证管理页/二维码链接一致）
HOST = os.environ.get("HOST", "0.0.0.0").strip() or "0.0.0.0"
PORT = int(os.environ.get("PORT", "8000").strip() or "8000")
# 默认简历保存目录（.env 里可覆盖为绝对路径；空值则用项目内默认）
_env_save_dir = os.environ.get("SAVE_DIR", "").strip()
DEFAULT_DIR = Path(_env_save_dir) if _env_save_dir else Path(DATA_DIR / "resumes")

app = FastAPI(title="BR TalentHub", version="2.0.0")

# 应用启动时确保数据表存在
db.init_db()

# 局域网现场部署，放开跨域让手机端任何来源都可访问
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")

ALLOWED_EXT = {".pdf", ".doc", ".docx"}
MAX_FILE_SIZE = 20 * 1024 * 1024  # 20MB


def now_str() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def gen_event_id() -> str:
    """短 ID：给每个招聘会一个唯一标识，二维码里携带。"""
    return uuid.uuid4().hex[:6].upper()


def get_local_ip() -> str:
    """探测本机局域网 IP（用于生成手机可访问的二维码链接）。"""
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.settimeout(1)
        s.connect(("8.8.8.8", 80))  # UDP 不真正发包，仅用于取本机路由 IP
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return "127.0.0.1"


def build_qr(url: str) -> Response:
    """生成二维码 PNG。"""
    qr = qrcode.QRCode(border=2, box_size=10, error_correction=qrcode.constants.ERROR_CORRECT_M)
    qr.add_data(url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buf = BytesIO()
    img.save(buf, format="PNG")
    return Response(content=buf.getvalue(), media_type="image/png")


# ---------------------------------------------------------------- 配置与招聘会

@app.get("/api/config")
def get_config():
    """管理端/上传页共用：当前招聘会配置 + 局域网地址 + 统计。"""
    event_name = db.get_setting("event_name", "BR 招聘会")
    event_id = db.get_setting("event_id", gen_event_id())
    if not db.get_setting("event_id"):
        db.set_setting("event_id", event_id)
    save_dir = db.get_setting("save_dir", str(DEFAULT_DIR))
    active_school = db.get_setting("active_school", "")
    return {
        "event_name": event_name,
        "event_id": event_id,
        "save_dir": save_dir,
        "host_ip": get_local_ip(),
        "port": PORT,
        "count": db.count_resumes(),
        "active_school": active_school,
        "updated_at": db.get_setting("updated_at", ""),
    }


@app.post("/api/config")
def update_config(payload: dict):
    """设置招聘会名称、保存目录、开始收集时间。"""
    event_name = (payload.get("event_name") or "").strip()
    save_dir = (payload.get("save_dir") or "").strip()

    if event_name:
        db.set_setting("event_name", event_name)
    if save_dir:
        path = Path(save_dir)
        try:
            path.mkdir(parents=True, exist_ok=True)
        except Exception as exc:
            raise HTTPException(400, f"无法创建保存目录：{exc}")
        db.set_setting("save_dir", str(path.resolve()))

    if "started_at" in payload:
        db.set_setting("started_at", payload["started_at"])
    db.set_setting("updated_at", now_str())
    return get_config()


@app.get("/api/stats")
def get_stats():
    """大屏实时统计：总数 + 最近上传记录。"""
    return {
        "count": db.count_resumes(),
        "recent": db.list_resumes(limit=20),
    }


@app.get("/api/qrcode")
def qr_code(school: str = ""):
    """生成二维码 PNG：指向手机上传页（携带招聘会 event_id 与学校）。

    二维码 URL：http://本机IP:{PORT}/upload?event={event_id}&school={校名}
    school 为空时生成通用码（不带学校，手机端可自选学校）。
    """
    cfg = get_config()
    base = f"http://{cfg['host_ip']}:{cfg['port']}"
    url = f"{base}/upload?event={cfg['event_id']}"
    if school:
        url += f"&school={school}"
    return build_qr(url)


@app.get("/api/schools/{school_id}/qrcode")
def school_qr_code(school_id: int):
    """某个学校的专属二维码。"""
    school = db.get_school(school_id)
    if not school:
        raise HTTPException(404, "学校不存在")
    cfg = get_config()
    base = f"http://{cfg['host_ip']}:{cfg['port']}"
    url = f"{base}/upload?event={cfg['event_id']}&school={school['name']}"
    return build_qr(url)


@app.post("/api/event/reset")
def reset_event():
    """开启一场新招聘会：换新 event_id + 清空统计。"""
    db.set_setting("event_id", gen_event_id())
    db.set_setting("started_at", now_str())
    db.set_setting("updated_at", now_str())
    return get_config()


# ---------------------------------------------------------------- 岗位管理

@app.get("/api/positions")
def list_positions():
    return db.list_positions()


@app.post("/api/positions")
def create_position(payload: dict):
    name = (payload.get("name") or "").strip()
    requirement = (payload.get("requirement") or "").strip()
    if not name:
        raise HTTPException(400, "岗位名称不能为空")
    if db.position_name_exists(name):
        raise HTTPException(400, f"岗位「{name}」已存在")
    position_id = db.add_position(name, requirement)
    return {"id": position_id, "name": name, "requirement": requirement}


@app.delete("/api/positions/{position_id}")
def delete_position(position_id: int):
    if not db.delete_position(position_id):
        raise HTTPException(404, "岗位不存在")
    return {"ok": True}


@app.put("/api/positions/{position_id}")
def update_position(position_id: int, payload: dict):
    """编辑岗位：改名称 / 岗位要求。"""
    name = (payload.get("name") or "").strip()
    requirement = (payload.get("requirement") or "").strip()
    if not name:
        raise HTTPException(400, "岗位名称不能为空")
    if db.position_name_exists(name, exclude_id=position_id):
        raise HTTPException(400, f"岗位「{name}」已存在")
    if not db.update_position(position_id, name, requirement):
        raise HTTPException(404, "岗位不存在")
    p = db.get_position(position_id)
    return {"id": p["id"], "name": p["name"], "requirement": p["requirement"]}


@app.post("/api/positions/import")
async def import_positions(file: UploadFile = File(...)):
    """Excel 导入岗位：识别"岗位名称""岗位要求"表头（支持别名），逐行创建。"""
    if not (file.filename or "").lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(400, "仅支持 .xlsx 格式的 Excel 文件")

    content = await file.read()
    if not content:
        raise HTTPException(400, "文件内容为空")
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(400, f"无法解析 Excel 文件：{exc}")
    ws = wb.active

    rows = ws.iter_rows(values_only=True)
    # 第一行表头
    header = next(rows, None)
    if not header:
        raise HTTPException(400, "Excel 为空（无表头）")

    # 识别"岗位名称""岗位要求"列（支持别名）
    name_idx = req_idx = None
    aliases = {
        "岗位名称": ["岗位名称", "岗位", "职位名称", "职位", "名称", "岗位名"],
        "岗位要求": ["岗位要求", "要求", "职位要求", "任职要求", "岗位描述", "职责"],
    }
    for i, cell in enumerate(header):
        text = str(cell or "").strip()
        if name_idx is None and text in aliases["岗位名称"]:
            name_idx = i
        if req_idx is None and text in aliases["岗位要求"]:
            req_idx = i
    if name_idx is None:
        raise HTTPException(400, "未识别到「岗位名称」列（请确认表头含 岗位名称/岗位/职位 等）")

    created, skipped, dup = 0, 0, []
    for row in rows:
        name = str(row[name_idx] or "").strip() if name_idx < len(row) else ""
        if not name:
            skipped += 1
            continue
        if db.position_name_exists(name):
            dup.append(name)
            continue
        requirement = str(row[req_idx] or "").strip() if req_idx is not None and req_idx < len(row) else ""
        db.add_position(name, requirement)
        created += 1
    return {
        "ok": True,
        "created": created,
        "skipped": skipped,
        "duplicates": dup,
    }


# ---------------------------------------------------------------- 学校管理

def _school_out(school: dict) -> dict:
    """学校记录 → 前端展示结构（把绑定岗位 JSON 展开为详情）。"""
    pos_ids = json_loads(school.get("positions") or "[]")
    positions = []
    for pid in pos_ids:
        p = db.get_position(pid)
        if p:
            positions.append({"id": p["id"], "name": p["name"]})
    return {
        "id": school["id"],
        "name": school["name"],
        "position_ids": pos_ids,
        "positions": positions,
        "created_at": school["created_at"],
    }


def json_loads(s: str) -> list:
    import json

    try:
        return json.loads(s)
    except Exception:
        return []


@app.get("/api/schools")
def list_schools():
    return [_school_out(s) for s in db.list_schools()]


@app.post("/api/schools")
def create_school(payload: dict):
    name = (payload.get("name") or "").strip()
    position_ids = payload.get("positions") or []
    if not name:
        raise HTTPException(400, "学校名称不能为空")
    if db.school_name_exists(name):
        raise HTTPException(400, f"学校「{name}」已存在")
    # 校验岗位 ID 存在
    valid = []
    for pid in position_ids:
        if db.get_position(pid):
            valid.append(pid)
    school_id = db.add_school(name, valid)
    return _school_out(db.get_school(school_id))


@app.put("/api/schools/{school_id}/positions")
def update_school_positions(school_id: int, payload: dict):
    school = db.get_school(school_id)
    if not school:
        raise HTTPException(404, "学校不存在")
    position_ids = payload.get("positions") or []
    valid = [pid for pid in position_ids if db.get_position(pid)]
    db.update_school_positions(school_id, valid)
    return _school_out(db.get_school(school_id))


@app.delete("/api/schools/{school_id}")
def delete_school(school_id: int):
    if not db.delete_school(school_id):
        raise HTTPException(404, "学校不存在")
    return {"ok": True}


@app.post("/api/schools/{school_id}/activate")
def activate_school(school_id: int):
    """将某学校设为当前招聘会学校（现场大屏切换用）。"""
    school = db.get_school(school_id)
    if not school:
        raise HTTPException(404, "学校不存在")
    db.set_setting("active_school", school["name"])
    db.set_setting("updated_at", now_str())
    return {"ok": True, "school": school["name"]}


@app.get("/api/options")
def get_options(school: str = ""):
    """手机端下拉数据源：学校列表 + 指定学校的岗位列表。"""
    schools = [_school_out(s) for s in db.list_schools()]
    positions = []
    if school:
        s = db.get_school_by_name(school)
        if s:
            positions = _school_out(s)["positions"]
    return {
        "schools": [{"id": s["id"], "name": s["name"]} for s in schools],
        "positions": positions,
        "school": school,
    }


# ---------------------------------------------------------------- 数据看板

@app.get("/api/dashboard")
def get_dashboard():
    """看板统计：总数 / 按学校 / 按岗位 / 近7日。"""
    stats = db.dashboard_stats()
    # 补充学校/岗位名称（供看板横向图用）
    return stats


# ---------------------------------------------------------------- 简历上传

@app.post("/api/resumes/upload")
async def upload_resume(
    file: UploadFile = File(...),
    name: str = Form(""),
    phone: str = Form(""),
    position: str = Form(""),
    school: str = Form(""),
    position_id: int = Form(0),
):
    """手机端提交简历：校验（姓名/手机/岗位必填）→ 按 学校/岗位/文件 三级目录落盘 → 记录。"""
    name = (name or "").strip()
    phone = (phone or "").strip()
    position = (position or "").strip()
    school = (school or "").strip()

    # 姓名 / 手机 / 岗位必填（与手机端前端校验一致）
    if not name:
        raise HTTPException(400, "请填写姓名")
    if not phone:
        raise HTTPException(400, "请填写手机号")
    if not position:
        raise HTTPException(400, "请选择应聘岗位")

    # 解析学校（二维码带 school 参数则锁定，否则取第一个学校）
    school_row = db.get_school_by_name(school) if school else None
    if not school_row:
        schools = db.list_schools()
        if not schools:
            raise HTTPException(400, "暂未配置招聘学校，请联系现场工作人员")
        school_row = schools[0]
        school = school_row["name"]
    school_id = school_row["id"]

    # 岗位：优先用 position_id（来自下拉框），否则按名称匹配
    position_row = db.get_position(position_id) if position_id else None
    if not position_row:
        for p in db.list_positions():
            if p["name"] == position:
                position_row = p
                break
    if not position_row:
        raise HTTPException(400, f"岗位「{position}」不存在")
    position_id = position_row["id"]

    # 学校-岗位归属校验：岗位必须属于该学校绑定列表（手机端下拉只显示该校岗位）
    school_pos_ids = json_loads(school_row.get("positions") or "[]")
    if position_id not in school_pos_ids:
        raise HTTPException(400, f"岗位「{position_row['name']}」不属于学校「{school_row['name']}」")

    original = file.filename or "resume"
    ext = Path(original).suffix.lower()
    if ext not in ALLOWED_EXT:
        raise HTTPException(400, "仅支持 PDF / DOC / DOCX 格式")

    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(400, "文件超过 20MB 上限")

    # 清理文件名里的不安全字符
    safe_original = re.sub(r'[\\/:*?"<>|\r\n]', "_", original)

    # 三级目录：保存目录 / 学校名 / 岗位名 / 文件
    base_dir = Path(db.get_setting("save_dir", str(DEFAULT_DIR)))
    safe_school = re.sub(r'[\\/:*?"<>|\r\n]', "_", school)
    safe_position = re.sub(r'[\\/:*?"<>|\r\n]', "_", position)
    target_dir = base_dir / safe_school / safe_position
    try:
        target_dir.mkdir(parents=True, exist_ok=True)
    except Exception as exc:
        raise HTTPException(400, f"无法创建存储目录：{exc}")

    ts = datetime.now().strftime("%Y%m%d%H%M%S")
    stored_name = f"{ts}_{uuid.uuid4().hex[:6]}{ext}"
    filepath = target_dir / stored_name
    filepath.write_bytes(content)

    record = {
        "name": name,
        "phone": phone,
        "position": position,
        "original": safe_original,
        "filename": stored_name,
        "filepath": str(filepath),
        "filesize": len(content),
        "upload_time": now_str(),
        "ip": "",
        "school_id": school_id,
        "position_id": position_id,
    }
    resume_id = db.add_resume(record)
    return {"id": resume_id, "message": "上传成功", "filename": original}


@app.get("/api/resumes")
def list_resumes(
    school: str = "",
    position: str = "",
    keyword: str = "",
    date_start: str = "",
    date_end: str = "",
    limit: int = 100,
    offset: int = 0,
):
    """简历管理：按 学校/岗位/关键词/日期时间段 筛选，分页。"""
    limit = min(max(limit, 1), 500)
    offset = max(offset, 0)
    return db.query_resumes(
        school=school,
        position=position,
        keyword=keyword,
        date_start=date_start,
        date_end=date_end,
        limit=limit,
        offset=offset,
    )


@app.delete("/api/resumes/{resume_id}")
def delete_resume(resume_id: int):
    """删除单条简历记录（同时删除磁盘文件）。"""
    row = db.get_resume(resume_id)
    if not row:
        raise HTTPException(404, "记录不存在")
    try:
        Path(row["filepath"]).unlink(missing_ok=True)
    except Exception:
        pass
    db.delete_resume(resume_id)
    return {"ok": True}


@app.get("/api/resumes/export.zip")
def export_resumes(
    school: str = "",
    position: str = "",
    keyword: str = "",
    date_start: str = "",
    date_end: str = "",
    ids: str = "",
):
    """按当前筛选条件（或指定 ids）打包下载简历文件（ZIP）。

    文件名保留 学校/岗位 目录结构，便于会后归档。
    """
    if ids:
        # 指定记录：按 id 导出
        id_list = [int(x) for x in ids.split(",") if x.strip().isdigit()]
        items = []
        for rid in id_list:
            row = db.get_resume(rid)
            if row:
                items.append(row)
    else:
        data = db.query_resumes(
            school=school,
            position=position,
            keyword=keyword,
            date_start=date_start,
            date_end=date_end,
            limit=500,
            offset=0,
        )
        items = data["items"]
    if not items:
        raise HTTPException(400, "没有符合条件的简历")

    buf = BytesIO()
    # 处理 zip 内文件名重复（同名原始文件名 + 落盘时间戳）
    seen = {}
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for r in items:
            path = Path(r["filepath"])
            if not path.exists():
                continue
            base = Path(r["original"]).stem
            ext = Path(r["original"]).suffix
            folder = f"{r.get('school_name') or '未分类'}/{r.get('position_name') or '未分类'}"
            name = base + ext
            if name in seen:
                seen[name] += 1
                name = f"{base}_{seen[name]}{ext}"
            else:
                seen[name] = 0
            zf.write(path, arcname=f"{folder}/{name}")
    buf.seek(0)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return Response(
        content=buf.getvalue(),
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="resumes_{stamp}.zip"'},
    )


@app.get("/api/resumes/{resume_id}/download")
def download_resume(resume_id: int):
    row = db.get_resume(resume_id)
    if not row:
        raise HTTPException(404, "记录不存在")
    path = Path(row["filepath"])
    if not path.exists():
        raise HTTPException(404, "文件已丢失")
    return FileResponse(path, filename=row["original"])


# ---------------------------------------------------------------- 页面路由

@app.get("/")
def index():
    return FileResponse(STATIC_DIR / "admin.html")


@app.get("/admin")
def admin():
    return FileResponse(STATIC_DIR / "admin.html")


@app.get("/upload")
def upload_page():
    return FileResponse(STATIC_DIR / "upload.html")


if __name__ == "__main__":
    db.init_db()
    uvicorn.run(app, host=HOST, port=PORT)
