"""岗位管理：/api/positions*（CRUD + Excel 导入）。"""
import io

from fastapi import APIRouter, File, HTTPException, UploadFile
from openpyxl import load_workbook

from app import database as db
from app.schemas import PositionPayload

router = APIRouter()


@router.get("/api/positions")
def list_positions():
    return db.list_positions()


@router.post("/api/positions")
def create_position(payload: PositionPayload):
    name = payload.name.strip()
    requirement = (payload.requirement or "").strip()
    if db.position_name_exists(name):
        raise HTTPException(400, f"岗位「{name}」已存在")
    position_id = db.add_position(name, requirement)
    return {"id": position_id, "name": name, "requirement": requirement}


@router.put("/api/positions/{position_id}")
def update_position(position_id: int, payload: PositionPayload):
    """编辑岗位：改名称 / 岗位要求。"""
    name = payload.name.strip()
    requirement = (payload.requirement or "").strip()
    if db.position_name_exists(name, exclude_id=position_id):
        raise HTTPException(400, f"岗位「{name}」已存在")
    if not db.update_position(position_id, name, requirement):
        raise HTTPException(404, "岗位不存在")
    p = db.get_position(position_id)
    return {"id": p["id"], "name": p["name"], "requirement": p["requirement"]}


@router.delete("/api/positions/{position_id}")
def delete_position(position_id: int):
    if not db.delete_position(position_id):
        raise HTTPException(404, "岗位不存在")
    return {"ok": True}


@router.post("/api/positions/import")
async def import_positions(file: UploadFile = File(...)):
    """Excel 导入岗位：识别"岗位名称""岗位要求"表头（支持别名），逐行创建。"""
    if not (file.filename or "").lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(400, "仅支持 .xlsx 格式的 Excel 文件")

    content = await file.read()
    if not content:
        raise HTTPException(400, "文件内容为空")
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(400, f"无法解析 Excel 文件：{exc}")
    ws = wb.active

    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        raise HTTPException(400, "Excel 为空（无表头）")

    # 识别"岗位名称""岗位要求"列（支持别名）
    name_idx = req_idx = None
    aliases = {
        "岗位名称": ["岗位名称", "岗位", "职位名称", "职位", "名称", "岗位名"],
        "岗位要求": ["岗位要求", "要求", "职位要求", "任职要求", "岗位描述", "职责"],
    }
    for i, cell in enumerate(header):
        text = str(cell or "").strip()
        if name_idx is None and text in aliases["岗位名称"]:
            name_idx = i
        if req_idx is None and text in aliases["岗位要求"]:
            req_idx = i
    if name_idx is None:
        raise HTTPException(400, "未识别到「岗位名称」列（请确认表头含 岗位名称/岗位/职位 等）")

    created, skipped, dup = 0, 0, []
    for row in rows:
        name = str(row[name_idx] or "").strip() if name_idx < len(row) else ""
        if not name:
            skipped += 1
            continue
        if db.position_name_exists(name):
            dup.append(name)
            continue
        requirement = str(row[req_idx] or "").strip() if req_idx is not None and req_idx < len(row) else ""
        db.add_position(name, requirement)
        created += 1
    return {
        "ok": True,
        "created": created,
        "skipped": skipped,
        "duplicates": dup,
    }
